#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_orbat.py — lecture / création / édition des ORBAT au format xlsx ODIN
(TRADOC G2), l'équivalent tableur du XML JTDS OBSV4.2.

Grammaire vérifiée à 100 % contre le XML de référence JRTC ARIANA-17 IFV DIV
(3160 unités, 44010 entités, 5838 remorquées, 29195 personnels, 19474 crew,
9721 passagers — tous les compteurs correspondent exactement après expansion).
Voir references/xlsx_format.md.

Modèle : un classeur = 1 feuille « UNIT INFO » (l'ossature) + N feuilles
« templates » (la composition interne réutilisable). Un ORBAT complet est
obtenu en *instanciant* chaque ligne T de UNIT INFO avec la feuille template
qu'elle nomme.

API principale
--------------
    wb   = load(path)                 -> Workbook (modèle brut, fidèle)
    tree = expand(wb)                 -> arbre d'unités instancié
    agg_equipment(tree) / agg_personnel(tree)
    save(wb, path)
    to_mapper_json(wb, ...)           -> dict ORBAT orbat-mapper.app
    validate(wb)                      -> liste de problèmes

CLI
---
    python xlsx_orbat.py read     F.xlsx
    python xlsx_orbat.py tree     F.xlsx [--unit NOM] [--depth N]
    python xlsx_orbat.py agg      F.xlsx [--unit NOM] [--what equipment|personnel]
    python xlsx_orbat.py validate F.xlsx
    python xlsx_orbat.py tojson   F.xlsx SORTIE.json [--side NOM] [--identity 3]
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl requis : pip install openpyxl")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

INFO_SHEET = "UNIT INFO"
INFO_HEADER = ["TYPE", "NAME", "PARENT", "UIC", "ECHELON",
               "UNIT CLASS", "TEMPLATE", "2525C"]
ROOT = "TOP"

# Types de ligne
#   UNIT INFO : U = unité conteneur (enfants = autres lignes UNIT INFO)
#               T = unité templatée (composition = feuille nommée en TEMPLATE)
#   template  : U = unité (sous-unité du template ; la 1re, parent TOP, EST
#                   l'unité qui référence le template)
#               E = plateforme / véhicule      (parent = NOM d'unité)
#               W = remorqué (trailer)         (parent = ID de l'E tracteur)
#               M = monté / emporté (ex. UAV)  (parent = ID de l'E porteur)
#               P = personnel                  (parent = ID de l'E, rôle C|P)
KIND_ENTITY = ("E", "W", "M")
ROLE_CREW, ROLE_PAX = "C", "P"


# ---------------------------------------------------------------------------
# Modèle
# ---------------------------------------------------------------------------
class Workbook:
    """Modèle fidèle du classeur : lignes brutes, sans interprétation."""

    def __init__(self, info=None, templates=None):
        self.info = info if info is not None else []      # list[dict]
        self.templates = templates if templates is not None else {}  # name -> list[dict]

    # --- accès pratiques ---
    def unit(self, name):
        for r in self.info:
            if r["name"] == name:
                return r
        return None

    def roots(self):
        return [r for r in self.info if (r["parent"] or ROOT) == ROOT]

    def children_of(self, name):
        return [r for r in self.info if r["parent"] == name]

    def __repr__(self):
        return (f"<Workbook {len(self.info)} unités UNIT INFO, "
                f"{len(self.templates)} templates>")


class Unit:
    """Unité instanciée (après expansion)."""

    __slots__ = ("name", "uic", "echelon", "unit_class", "sidc2525",
                 "template", "parent", "children", "entities", "origin")

    def __init__(self, name, echelon=None, unit_class=None, sidc2525=None,
                 uic=None, template=None, origin=None):
        self.name = name
        self.uic = uic
        self.echelon = echelon
        self.unit_class = unit_class
        self.sidc2525 = sidc2525
        self.template = template
        self.origin = origin          # "info" | nom du template
        self.parent = None
        self.children = []
        self.entities = []            # list[Entity]

    def walk(self):
        yield self
        for c in self.children:
            yield from c.walk()

    def __repr__(self):
        return f"<Unit {self.name!r} {self.echelon} {len(self.entities)}ent {len(self.children)}sub>"


class Entity:
    """Plateforme (E), remorque (W), emport (M) ou personnel (P) instancié."""

    __slots__ = ("kind", "cls", "role", "carrier", "crew", "pax", "unit")

    def __init__(self, kind, cls, role=None):
        self.kind = kind              # E | W | M | P
        self.cls = cls
        self.role = role              # pour P : C (crew) ou P (passager)
        self.carrier = None           # Entity porteuse (pour W/M/P)
        self.crew = []
        self.pax = []
        self.unit = None

    def __repr__(self):
        return f"<{self.kind} {self.cls!r}>"


# ---------------------------------------------------------------------------
# Lecture
# ---------------------------------------------------------------------------
def _cell(row, i):
    return row[i] if i < len(row) else None


def load(path):
    """Charge un classeur ORBAT ODIN en modèle fidèle."""
    xl = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if INFO_SHEET not in xl.sheetnames:
        raise ValueError(
            f"feuille « {INFO_SHEET} » absente — ce classeur n'est pas un ORBAT "
            f"ODIN (feuilles trouvées : {xl.sheetnames[:5]}…)")

    wb = Workbook()

    ws = xl[INFO_SHEET]
    for n, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row or not row[0]:
            continue
        if n == 1 and str(row[0]).strip().upper() == "TYPE":
            continue                                    # en-tête
        wb.info.append({
            "type": str(row[0]).strip().upper(),
            "name": _cell(row, 1),
            "parent": _cell(row, 2),
            "uic": _cell(row, 3),
            "echelon": _cell(row, 4),
            "unit_class": _cell(row, 5),
            "template": _cell(row, 6),
            "sidc2525": _cell(row, 7),
            "_row": n,
        })

    for name in xl.sheetnames:
        if name == INFO_SHEET:
            continue
        rows = []
        for n, row in enumerate(xl[name].iter_rows(values_only=True), 1):
            if not row or not row[0]:
                continue
            k = str(row[0]).strip().upper()
            if k == "U":
                rows.append({"kind": "U", "name": _cell(row, 1),
                             "parent": _cell(row, 2), "uic": _cell(row, 3),
                             "echelon": _cell(row, 4), "unit_class": _cell(row, 5),
                             "sidc2525": _cell(row, 6), "_row": n})
            elif k in KIND_ENTITY:
                rows.append({"kind": k, "id": _cell(row, 1),
                             "parent": _cell(row, 2), "cls": _cell(row, 3),
                             "_row": n})
            elif k == "P":
                rows.append({"kind": "P", "id": _cell(row, 1),
                             "parent": _cell(row, 2), "role": _cell(row, 3),
                             "cls": _cell(row, 4), "_row": n})
            else:
                rows.append({"kind": k, "_raw": list(row), "_row": n})
        wb.templates[name] = rows

    xl.close()
    return wb


# ---------------------------------------------------------------------------
# Expansion (instanciation des templates)
# ---------------------------------------------------------------------------
def _instantiate(tpl_rows, host, template_name):
    """
    Peuple `host` (Unit) avec le contenu d'une feuille template.
    La 1re ligne U (parent TOP) EST `host` : ses attributs ne sont pas
    ré-appliqués (UNIT INFO fait autorité), seuls ses enfants/entités le sont.

    ⚠ RÉSOLUTION SÉQUENTIELLE — les noms d'unités NE SONT PAS uniques dans une
    feuille (ex. « SPT TM 1 » existe sous HQ 1 CO *et* sous HQ TANK CO dans
    IFV BN). La feuille se lit de haut en bas : une référence par nom (parent
    d'un U, ou parent d'un E) désigne la **définition U la plus récente**
    portant ce nom. Vérifié contre le XML : « SPT TM 1 » sous HQ 1 CO a 0
    entité, celui sous HQ TANK CO en a 24. Un parseur qui indexe par nom
    fusionne silencieusement les homonymes et sur-compte le matériel.
    """
    current = {}      # nom -> Unit la plus récemment déclarée
    by_id = {}        # id -> Entity
    root_seen = False

    for r in tpl_rows:
        k = r["kind"]

        if k == "U":
            if not root_seen and (r["parent"] or ROOT) == ROOT:
                root_seen = True
                current[r["name"]] = host
                continue
            parent = current.get(r["parent"])
            if parent is None:
                raise ValueError(
                    f"template « {template_name} » ligne {r['_row']} : unité "
                    f"« {r['name']} » référence le parent « {r['parent']} », "
                    "non déclaré plus haut dans la feuille")
            u = Unit(r["name"], echelon=r["echelon"], unit_class=r["unit_class"],
                     sidc2525=r["sidc2525"], uic=r["uic"], origin=template_name)
            u.parent = parent
            parent.children.append(u)
            current[r["name"]] = u

        elif k == "E":
            u = current.get(r["parent"])
            if u is None:
                raise ValueError(
                    f"template « {template_name} » ligne {r['_row']} : E "
                    f"rattaché à « {r['parent']} », non déclaré plus haut")
            e = Entity("E", r["cls"])
            e.unit = u
            u.entities.append(e)
            by_id[r["id"]] = e

        elif k in ("W", "M"):
            carrier = by_id.get(r["parent"])
            if carrier is None:
                raise ValueError(
                    f"template « {template_name} » ligne {r['_row']} : {k} "
                    f"référence le porteur id={r['parent']!r}, non défini plus haut")
            e = Entity(k, r["cls"])
            e.carrier = carrier
            e.unit = carrier.unit
            carrier.unit.entities.append(e)
            by_id[r["id"]] = e

        elif k == "P":
            carrier = by_id.get(r["parent"])
            if carrier is None:
                raise ValueError(
                    f"template « {template_name} » ligne {r['_row']} : P "
                    f"référence la plateforme id={r['parent']!r}, non définie plus haut")
            p = Entity("P", r["cls"], role=r["role"])
            p.carrier = carrier
            p.unit = carrier.unit
            (carrier.crew if r["role"] == ROLE_CREW else carrier.pax).append(p)
            carrier.unit.entities.append(p)
            by_id[r["id"]] = p

    if not root_seen:
        raise ValueError(f"template « {template_name} » : aucune ligne U racine (parent {ROOT})")
    return host


def expand(wb, root=None):
    """
    Instancie l'ORBAT complet. Renvoie la (ou les) Unit racine(s).
    Si `root` est donné, ne renvoie que cette unité (et son sous-arbre).
    """
    units = {}
    for r in wb.info:
        u = Unit(r["name"], echelon=r["echelon"], unit_class=r["unit_class"],
                 sidc2525=r["sidc2525"], uic=r["uic"], template=r["template"],
                 origin="info")
        units[r["name"]] = u

    roots = []
    for r in wb.info:
        u = units[r["name"]]
        p = r["parent"]
        if not p or p == ROOT:
            roots.append(u)
        else:
            parent = units.get(p)
            if parent is None:
                raise ValueError(
                    f"UNIT INFO ligne {r['_row']} : « {r['name']} » référence "
                    f"le parent inconnu « {p} »")
            u.parent = parent
            parent.children.append(u)

    # instanciation des templates (lignes T uniquement ; pour les U, la colonne
    # TEMPLATE n'est qu'une étiquette doctrinale, jamais une feuille à étendre)
    for r in wb.info:
        if r["type"] != "T":
            continue
        tpl = wb.templates.get(r["template"])
        if tpl is None:
            raise ValueError(
                f"UNIT INFO ligne {r['_row']} : « {r['name']} » est de type T "
                f"mais la feuille template « {r['template']} » n'existe pas")
        _instantiate(tpl, units[r["name"]], r["template"])

    if root is not None:
        u = units.get(root)
        if u is None:
            raise KeyError(f"unité « {root} » absente de UNIT INFO")
        return u
    return roots[0] if len(roots) == 1 else roots


# ---------------------------------------------------------------------------
# Agrégation
# ---------------------------------------------------------------------------
def _iter(tree):
    if isinstance(tree, Unit):
        yield from tree.walk()
    else:
        for t in tree:
            yield from t.walk()


def agg_equipment(tree):
    """Compte les plateformes/remorques/emports (E, W, M) par classe."""
    c = Counter()
    for u in _iter(tree):
        for e in u.entities:
            if e.kind in KIND_ENTITY:
                c[e.cls] += 1
    return c


def agg_personnel(tree):
    """Compte les personnels (P) par classe."""
    c = Counter()
    for u in _iter(tree):
        for e in u.entities:
            if e.kind == "P":
                c[e.cls] += 1
    return c


def counts(tree):
    """Compteurs globaux, comparables aux totaux du XML JTDS."""
    n = Counter()
    for u in _iter(tree):
        n["units"] += 1
        for e in u.entities:
            n["entities"] += 1
            n[e.kind] += 1
            if e.kind == "P":
                n["crew" if e.role == ROLE_CREW else "pax"] += 1
    n["towed"] = n["W"] + n["M"]
    return n


# ---------------------------------------------------------------------------
# Écriture
# ---------------------------------------------------------------------------
def save(wb, path):
    """Écrit le modèle en xlsx (grammaire ODIN, police Arial)."""
    from openpyxl.styles import Font
    xl = openpyxl.Workbook()
    ws = xl.active
    ws.title = INFO_SHEET
    ws.append(INFO_HEADER)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True)
    for r in wb.info:
        ws.append([r["type"], r["name"], r["parent"], r["uic"], r["echelon"],
                   r["unit_class"], r["template"], r["sidc2525"]])

    for name, rows in wb.templates.items():
        s = xl.create_sheet(title=name[:31])
        for r in rows:
            k = r["kind"]
            if k == "U":
                s.append(["U", r["name"], r["parent"], r["uic"], r["echelon"],
                          r["unit_class"], r["sidc2525"]])
            elif k in KIND_ENTITY:
                s.append([k, r["id"], r["parent"], r["cls"]])
            elif k == "P":
                s.append(["P", r["id"], r["parent"], r["role"], r["cls"]])
            else:
                s.append(r.get("_raw", [k]))

    for s in xl.worksheets:
        for row in s.iter_rows():
            for c in row:
                if c.font is None or c.font.name != "Arial":
                    c.font = Font(name="Arial", bold=c.font.bold if c.font else False)
    xl.save(path)
    return path


# ---------------------------------------------------------------------------
# Construction / édition
# ---------------------------------------------------------------------------
def new_workbook():
    return Workbook()


def add_unit(wb, name, parent=ROOT, echelon=None, unit_class=None,
             sidc2525=None, uic=None, template=None):
    """Ajoute une unité à UNIT INFO. type=T si un template est nommé, sinon U."""
    if wb.unit(name):
        raise ValueError(f"unité « {name} » déjà présente dans UNIT INFO")
    wb.info.append({
        "type": "T" if template else "U", "name": name, "parent": parent,
        "uic": uic, "echelon": echelon, "unit_class": unit_class,
        "template": template, "sidc2525": sidc2525,
        "_row": len(wb.info) + 2,
    })
    return wb.info[-1]


def remove_unit(wb, name, cascade=True):
    """Retire une unité de UNIT INFO (et ses descendants si cascade)."""
    victims = {name}
    if cascade:
        changed = True
        while changed:
            changed = False
            for r in wb.info:
                if r["parent"] in victims and r["name"] not in victims:
                    victims.add(r["name"])
                    changed = True
    before = len(wb.info)
    wb.info = [r for r in wb.info if r["name"] not in victims]
    return before - len(wb.info)


class TemplateBuilder:
    """
    Construit une feuille template en gérant la numérotation séquentielle des
    IDs (espace d'IDs unique et partagé entre E, W, M et P).

        tb = TemplateBuilder("IFV PLT", echelon="Platoon",
                             unit_class="MECH INF PLT", sidc2525="SNGP...")
        sqd = tb.unit("1 SQD", parent="IFV PLT", echelon="Squad", ...)
        v   = tb.platform("BMP 2 IFV", unit="1 SQD")
        tb.crew(v, "GL RFL AK74 5.45MM")
        tb.passenger(v, "GL RPG-29 VAMPIR 105MM")
        tb.towed(v, "TRL CGO .5T TO 2T")
        wb.templates["IFV PLT"] = tb.rows
    """

    def __init__(self, name, echelon=None, unit_class=None, sidc2525=None, uic=None):
        self.name = name
        self.rows = [{"kind": "U", "name": name, "parent": ROOT, "uic": uic,
                      "echelon": echelon, "unit_class": unit_class,
                      "sidc2525": sidc2525, "_row": 1}]
        self._next = 1

    def _id(self):
        i = self._next
        self._next += 1
        return i

    def unit(self, name, parent, echelon=None, unit_class=None, sidc2525=None, uic=None):
        self.rows.append({"kind": "U", "name": name, "parent": parent, "uic": uic,
                          "echelon": echelon, "unit_class": unit_class,
                          "sidc2525": sidc2525, "_row": len(self.rows) + 1})
        return name

    def platform(self, cls, unit):
        i = self._id()
        self.rows.append({"kind": "E", "id": i, "parent": unit, "cls": cls,
                          "_row": len(self.rows) + 1})
        return i

    def towed(self, carrier_id, cls):
        i = self._id()
        self.rows.append({"kind": "W", "id": i, "parent": carrier_id, "cls": cls,
                          "_row": len(self.rows) + 1})
        return i

    def mounted(self, carrier_id, cls):
        i = self._id()
        self.rows.append({"kind": "M", "id": i, "parent": carrier_id, "cls": cls,
                          "_row": len(self.rows) + 1})
        return i

    def _person(self, carrier_id, cls, role):
        i = self._id()
        self.rows.append({"kind": "P", "id": i, "parent": carrier_id,
                          "role": role, "cls": cls, "_row": len(self.rows) + 1})
        return i

    def crew(self, carrier_id, cls):
        return self._person(carrier_id, cls, ROLE_CREW)

    def passenger(self, carrier_id, cls):
        return self._person(carrier_id, cls, ROLE_PAX)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def validate(wb):
    """Contrôles d'intégrité référentielle. Renvoie list[str]."""
    errs = []
    names = [r["name"] for r in wb.info]
    for n, c in Counter(names).items():
        if c > 1:
            errs.append(f"UNIT INFO : nom d'unité dupliqué « {n} » ({c}×) — "
                        "les parents deviennent ambigus")
    known = set(names)
    roots = 0
    for r in wb.info:
        if r["type"] not in ("U", "T"):
            errs.append(f"UNIT INFO ligne {r['_row']} : TYPE « {r['type']} » "
                        "inconnu (attendu U ou T)")
        if not r["name"]:
            errs.append(f"UNIT INFO ligne {r['_row']} : NAME vide")
        p = r["parent"]
        if not p or p == ROOT:
            roots += 1
        elif p not in known:
            errs.append(f"UNIT INFO ligne {r['_row']} : « {r['name']} » a pour "
                        f"parent « {p} », absent de UNIT INFO")
        if r["type"] == "T":
            if not r["template"]:
                errs.append(f"UNIT INFO ligne {r['_row']} : « {r['name']} » est "
                            "de type T mais TEMPLATE est vide")
            elif r["template"] not in wb.templates:
                errs.append(f"UNIT INFO ligne {r['_row']} : « {r['name']} » "
                            f"référence la feuille « {r['template']} », absente")
        if r["sidc2525"] and len(str(r["sidc2525"])) != 15:
            errs.append(f"UNIT INFO ligne {r['_row']} : 2525C « {r['sidc2525']} » "
                        f"fait {len(str(r['sidc2525']))} car. (attendu 15)")
    if roots == 0:
        errs.append(f"UNIT INFO : aucune unité racine (parent « {ROOT} »)")

    used = {r["template"] for r in wb.info if r["type"] == "T"}
    for name, rows in wb.templates.items():
        if name not in used:
            errs.append(f"template « {name} » : jamais référencé par UNIT INFO "
                        "(feuille morte)")
        ids, tunits, rootc = set(), set(), 0
        for r in rows:
            if r["kind"] == "U":
                tunits.add(r["name"])
                if (r["parent"] or ROOT) == ROOT:
                    rootc += 1
            elif r["kind"] in KIND_ENTITY or r["kind"] == "P":
                if r["id"] in ids:
                    errs.append(f"template « {name} » ligne {r['_row']} : "
                                f"id {r['id']} dupliqué")
                ids.add(r["id"])
            else:
                errs.append(f"template « {name} » ligne {r['_row']} : type de "
                            f"ligne « {r['kind']} » inconnu (attendu U/E/W/M/P)")
        if rootc != 1:
            errs.append(f"template « {name} » : {rootc} ligne(s) U racine "
                        f"(parent {ROOT}) — il en faut exactement 1")
        seen = set()
        for r in rows:
            k = r["kind"]
            if k == "U" and r["name"] != next((x["name"] for x in rows
                                               if x["kind"] == "U"), None):
                if r["parent"] not in tunits:
                    errs.append(f"template « {name} » ligne {r['_row']} : parent "
                                f"« {r['parent']} » inconnu dans la feuille")
            elif k == "E":
                if r["parent"] not in tunits:
                    errs.append(f"template « {name} » ligne {r['_row']} : E "
                                f"rattaché à l'unité inconnue « {r['parent']} »")
            elif k in ("W", "M", "P"):
                if r["parent"] not in seen:
                    errs.append(f"template « {name} » ligne {r['_row']} : {k} "
                                f"référence l'id {r['parent']!r} non défini plus haut")
                if k == "P" and r["role"] not in (ROLE_CREW, ROLE_PAX):
                    errs.append(f"template « {name} » ligne {r['_row']} : rôle "
                                f"« {r['role']} » invalide (attendu C ou P)")
            if k in KIND_ENTITY or k == "P":
                seen.add(r["id"])
    return errs


# ---------------------------------------------------------------------------
# Conversion vers orbat-mapper.app
# ---------------------------------------------------------------------------
# Échelon ODIN (mot anglais) -> échelon APP-6D. Table sûre : le vocabulaire
# ODIN est en anglais US, donc « Section » = section US (13) et « Platoon » =
# peloton (14). ⚠ Ne pas confondre avec le français, où « Section » = 14.
ECHELON_MAP = {
    "Team": "11", "Crew": "11", "Squad": "12", "Section": "13",
    "Platoon": "14", "Detachment": "14", "Company": "15", "Battery": "15",
    "Troop": "15", "Battalion": "16", "Squadron": "16", "Regiment": "17",
    "Group": "17", "Brigade": "18", "Division": "21", "Corps": "22",
    "Army": "23", "Army Group": "24", "Command": "26",
}

# Affiliation 2525C (pos 2) -> identité APP-6D (pos 4).
AFFILIATION_MAP = {
    "P": "0", "U": "1", "A": "2", "F": "3", "N": "4", "S": "5", "H": "6",
    "G": "0", "W": "3", "D": "2", "L": "4", "M": "6", "J": "6", "K": "6",
}

# Famille de function ID 2525C (pos 5-10) -> entité APP-6D (pos 11-16).
#
# PROVENANCE — table dérivée, pas inventée :
#   [sidc.md] entités documentées dans references/sidc.md (conventions vérifiées
#             de Florian) — font autorité ;
#   [dict]    déduites par concordance des noms d'unités ODIN avec les exemples
#             de references/sidc_dictionary.json, en n'utilisant QUE les noms
#             non ambigus (un nom générique comme « HQ » ou « CMD SEC » apparaît
#             sous des dizaines de types et pollue le vote — il est écarté).
#
# Tout code absent d'ici est converti en entité générique 000000 ET listé dans
# le rapport de conversion. Les familles délibérément NON mappées faute de
# base fiable (à compléter par Florian plutôt qu'à deviner) :
#   UCA--- blindé/char · UCAWW- · UCDML- · UCII-- · USXOM- · UUMRS-
#   NB---- · NU---- (dimension non terrestre)
# Le dictionnaire ne contient pas d'entité « char » distincte de l'état-major
# (120500 = commandement), donc UCA--- reste volontairement non mappé.
FUNCTION_MAP = {
    # — manœuvre
    "UCA---": "120500",   # [dict] TANK CO/BN/BDE — blindés / cavalerie
    "UCAWW-": "120300",   # [APP-6D] AMPHIB PLT — amphibie (non couvert par le dict)
    "UCIZ--": "121000",   # [sidc.md] infanterie mécanisée
    "UCII--": "121105",   # [dict] HQ 17 IFV DIV / HQ 171 IFV BDE — mécanisée sur IFV
    "UCIL--": "121100",   # [sidc.md] infanterie légère
    "UCAA--": "120400",   # [sidc.md] anti-char
    "UCR---": "120501",   # [sidc.md] reconnaissance
    "UCRVA-": "120501",   # [sidc.md+dict] recon véhicule
    "UCRVM-": "120501",   # [dict] HM REC SEC / HM REC SQD
    "UCRX--": "121300",   # [dict] RECON PLT, LR RECON PLT, LR SENSOR PLT
    "UCV---": "120600",   # [sidc.md] AIR OPS SEC / AIR OPS TM
    "UCVU--": "121900",   # [dict] UAV PLT, TRANS LAUNCH SEC
    "GS----": "121700",   # [dict] SOF SEC, SOF TM
    "GC----": "110200",   # [dict] CA SEC, CA CELL
    # — appui feu
    "UCF---": "130300",   # [dict] artillerie (entité dominante du dictionnaire)
    "UCFRM-": "130300",   # [dict] MRL — famille artillerie
    "UCFT--": "130300",   # [dict] FIRE CNTRL SEC, OBS SEC
    "UCFS--": "130900",   # [dict] SURVEY SEC
    "UCFO--": "130600",   # [dict] MET SURV PLT, MET SEC
    "USAM--": "161600",   # [dict] MORT SEC
    # — défense sol-air
    "UCD---": "130102",   # [dict] HQ SAM BTY, MANPAD BTY, CMD TM SAM
    "UCDM--": "130102",   # [dict] SAM BTY
    "UCDML-": "130102",   # [dict] 1..3 PLT SAM
    # — génie
    "UCE---": "140700",   # [dict] ENG TM
    "NB----": "140700",   # [dict] FERRY PLT, ASLT BOAT SQD, POWERBOAT SQD — franchissement
    "NU----": "140700",   # [dict] DIVING PLT — plongeurs du génie
    "UCEC--": "140700",   # [dict] CBT ENG PLT, 1..3 SQD CBT ENG
    "UCEN--": "140700",   # [dict] CONST PLT, RD BRIDGE CONST CO
    # — NBC
    "UUAC--": "140100",   # [dict] CHEM SEC, CHEM TM
    "UUACR-": "140100",   # [dict] NBC RECON SQD
    "UUAD--": "140100",   # [dict] VEH/PERS DECON SQD, DECON CO
    # — transmissions / renseignement / IO
    "UUS---": "111000",   # [dict] SIG SEC, SIG PLT, COUR SQD, DGT TM
    "UUSR--": "111001",   # [dict] RADIO SEC, 1..3 SEC RADIO
    "UUI---": "110600",   # [sidc.md] information / influence
    "UUM---": "151000",   # [dict] MI PLT, IO SEC, RECON MGT SEC
    "UUMS--": "150503",   # [dict] SIG RECON PLT, 1/2 SEC COLL
    "UUMRG-": "150503",   # [dict] RADAR RECON PLT
    "UUMSE-": "150500",   # [dict] IEW CO, EW SEC
    "UUMO--": "150700",   # [dict] INT TM IEW
    "UUMRS-": "151000",   # [famille UUM] noms génériques — rens. militaire
    "UULM--": "141200",   # [dict] FOR PRO SEC, TAC SCTY PLT
    # — soutien / logistique
    "US----": "160600",   # [dict] SPT CO, GEN SPT SEC
    "USS---": "160600",   # [dict] LOG SEC, SPT SEC, SPT TM
    "USAQ--": "160600",   # [dict] LOG SQD
    "USA---": "160100",   # [dict] ADMIN SEC, ADM DISPO SEC
    "USS1--": "163700",   # [dict] MESS SEC
    "USS3--": "163900",   # [dict] POL SEC, POL PLT (carburants)
    "USS5--": "164100",   # [dict] AMMO SPLY PLT
    "UST---": "163600",   # [dict] SPLY TM, HHC SPLY SQD (transport)
    "USM---": "161300",   # [dict] MED SEC, TREATMENT SEC, SURG PLT
    "USMD--": "161300",   # [dict] DENT SEC
    # — maintenance
    "USX---": "161100",   # [dict] MNT SEC, MNT SQD
    "USXE--": "161100",   # [dict] EO SEC
    "USXR--": "161100",   # [dict] RECOVERY PLT
    "USXO--": "162300",   # [dict] ORD MNT PLT, ORD MNT CO
    "USXOM-": "162300",   # [famille USXO] maintenance ordonnance / missiles
}
DEFAULT_ENTITY = "000000"


def _sidc_from(unit, identity=None, report=None):
    """Construit un sidc APP-6D 20 chiffres depuis les champs ODIN."""
    from new_orbat import build_sidc

    ech = ECHELON_MAP.get(str(unit.echelon or "").strip())
    if ech is None:
        if report is not None:
            report["echelons_inconnus"][str(unit.echelon)] += 1
        ech = "00"

    code = str(unit.sidc2525 or "")
    func = code[4:10] if len(code) == 15 else None
    ent = FUNCTION_MAP.get(func)
    if ent is None:
        if report is not None and func:
            report["functions_non_mappees"][func] += 1
        ent = DEFAULT_ENTITY

    ident = identity
    if ident is None:
        ident = AFFILIATION_MAP.get(code[1], "3") if len(code) == 15 else "3"

    return build_sidc(ech, ent, identity=ident)


def to_mapper_json(wb, side_name=None, identity=None, name=None):
    """
    Convertit un ORBAT ODIN xlsx en scénario orbat-mapper.app.

    Chaque unité ODIN devient une Unit orbat-mapper (avec sidc APP-6D) ;
    ses E/W/M deviennent des `equipment` et ses P des `personnel`, agrégés
    par classe au niveau de l'unité qui les porte. La structure crew/passager
    et le lien porteur→remorque ne sont PAS représentables dans le format
    orbat-mapper : ils sont perdus (documenté dans references/xlsx_format.md).

    Renvoie (orbat_dict, rapport).
    """
    from new_orbat import make_unit, make_side, make_group, new_orbat

    report = {"functions_non_mappees": Counter(), "echelons_inconnus": Counter()}
    tree = expand(wb)
    roots = tree if isinstance(tree, list) else [tree]

    eq_cat, pe_cat = Counter(), Counter()

    def conv(u):
        eq = Counter()
        pe = Counter()
        for e in u.entities:
            if e.kind in KIND_ENTITY:
                eq[e.cls] += 1
            else:
                pe[e.cls] += 1
        eq_cat.update(eq)
        pe_cat.update(pe)
        desc = []
        if u.uic:
            desc.append(f"UIC {u.uic}")
        if u.unit_class:
            desc.append(f"Unit class ODIN : {u.unit_class}")
        if u.sidc2525:
            desc.append(f"2525C d'origine : {u.sidc2525}")
        if u.echelon:
            desc.append(f"Échelon ODIN : {u.echelon}")
        kw = {}
        if desc:
            kw["description"] = " · ".join(desc)
        if eq:
            kw["equipment"] = sorted(eq.items())
        if pe:
            kw["personnel"] = sorted(pe.items())
        # subUnits : toujours des Unit récursives avec sidc (jamais de Group
        # imbriqué — cf. references/format.md).
        return make_unit(u.name, "00", "000000",
                         subUnits=[conv(c) for c in u.children],
                         sidc=_sidc_from(u, identity, report), **kw)

    units = [conv(r) for r in roots]
    side = make_side(side_name or "ODIN", [make_group(name or "ORBAT", units)])

    orbat = new_orbat(
        name or (roots[0].name if roots else "ORBAT ODIN"),
        [side],
        equipment=[(k, "Classe d'entité ODIN (plateforme / remorque / emport)")
                   for k in sorted(eq_cat)],
        personnel=[(k, "Classe de personnel ODIN (LifeForm)")
                   for k in sorted(pe_cat)],
        description=(
            "Converti depuis un ORBAT ODIN (TRADOC G2) au format xlsx. "
            "Les quantités sont agrégées par classe au niveau de l'unité "
            "porteuse ; la structure équipage/passager et les liens "
            "porteur→remorque du format ODIN ne sont pas représentables ici "
            "et ont été perdus à la conversion."
        ),
    )
    return orbat, report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _print_tree(u, depth, cur=0, prefix=""):
    if depth is not None and cur > depth:
        return
    eq = sum(1 for e in u.entities if e.kind in KIND_ENTITY)
    pe = sum(1 for e in u.entities if e.kind == "P")
    tag = f"  [{eq} eq / {pe} pers]" if (eq or pe) else ""
    src = f"  ({u.template})" if u.template else ""
    print(f"{prefix}{u.name}  — {u.echelon or '?'}{tag}{src}")
    for c in u.children:
        _print_tree(c, depth, cur + 1, prefix + "    ")


def main():
    ap = argparse.ArgumentParser(description="ORBAT xlsx ODIN — lire / vérifier / convertir")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("read", help="résumé du classeur")
    p.add_argument("xlsx")

    p = sub.add_parser("tree", help="arbre des unités")
    p.add_argument("xlsx")
    p.add_argument("--unit")
    p.add_argument("--depth", type=int, default=2)

    p = sub.add_parser("agg", help="somme du matériel / personnel")
    p.add_argument("xlsx")
    p.add_argument("--unit")
    p.add_argument("--what", choices=["equipment", "personnel"], default="equipment")
    p.add_argument("--top", type=int, default=0)

    p = sub.add_parser("validate", help="intégrité référentielle")
    p.add_argument("xlsx")

    p = sub.add_parser("tojson", help="conversion vers orbat-mapper.app")
    p.add_argument("xlsx")
    p.add_argument("out")
    p.add_argument("--side")
    p.add_argument("--identity")
    p.add_argument("--name")

    a = ap.parse_args()
    wb = load(a.xlsx)

    if a.cmd == "read":
        print(wb)
        t = Counter(r["type"] for r in wb.info)
        print(f"  UNIT INFO : {dict(t)}")
        print(f"  templates : {len(wb.templates)}")
        n = counts(expand(wb))
        print(f"\n  Après expansion :")
        print(f"    unités          : {n['units']}")
        print(f"    entités         : {n['entities']}")
        print(f"      plateformes E : {n['E']}")
        print(f"      remorquées W  : {n['W']}")
        print(f"      emports M     : {n['M']}")
        print(f"      personnels P  : {n['P']}  (crew {n['crew']} / passagers {n['pax']})")

    elif a.cmd == "tree":
        t = expand(wb, root=a.unit)
        for r in (t if isinstance(t, list) else [t]):
            _print_tree(r, a.depth)

    elif a.cmd == "agg":
        t = expand(wb, root=a.unit)
        c = agg_equipment(t) if a.what == "equipment" else agg_personnel(t)
        items = c.most_common(a.top or None)
        w = max((len(k or "") for k, _ in items), default=10)
        for k, v in items:
            print(f"  {str(k):<{w}}  {v:>6}")
        print(f"  {'—' * w}  {'—' * 6}")
        print(f"  {'TOTAL':<{w}}  {sum(c.values()):>6}")

    elif a.cmd == "validate":
        errs = validate(wb)
        try:
            n = counts(expand(wb))
            print(f"Expansion OK : {n['units']} unités, {n['entities']} entités.")
        except Exception as e:
            errs.append(f"expansion impossible : {e}")
        print(f"PROBLÈMES : {len(errs)}\n")
        for e in errs[:60]:
            print("  ", e)
        if len(errs) > 60:
            print(f"   … (+{len(errs) - 60} autres)")
        sys.exit(1 if errs else 0)

    elif a.cmd == "tojson":
        orbat, rep = to_mapper_json(wb, side_name=a.side, identity=a.identity, name=a.name)
        with open(a.out, "w", encoding="utf-8") as f:
            json.dump(orbat, f, ensure_ascii=False, indent=2)
        print(f"Écrit {a.out}")
        if rep["functions_non_mappees"]:
            print("\n⚠ function ID 2525C non mappés -> entité générique 000000 :")
            for k, v in rep["functions_non_mappees"].most_common():
                print(f"    {k}  ×{v}")
            print("  Compléter FUNCTION_MAP dans xlsx_orbat.py pour les traiter.")
        if rep["echelons_inconnus"]:
            print("\n⚠ échelons ODIN non reconnus :")
            for k, v in rep["echelons_inconnus"].most_common():
                print(f"    {k!r}  ×{v}")


if __name__ == "__main__":
    main()
